# -*- coding: utf-8 -*-
import io
import os
import re
from contextlib import contextmanager
from datetime import datetime

import pandas as pd
import streamlit as st
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from werkzeug.utils import secure_filename

from app import create_app
from app.excel_importer import import_excel_complete, rollback_upload as do_rollback
from app.models import db, Manager, Order, UploadHistory


UPLOAD_DIR = "uploads"


@st.cache_resource
def get_flask_app():
    return create_app()


@contextmanager
def app_context():
    app = get_flask_app()
    with app.app_context():
        yield


def extract_field(notes, field_name):
    if not notes:
        return ""
    pattern = rf"{field_name}:\s*([^,]+)"
    match = re.search(pattern, notes)
    return match.group(1).strip() if match else ""


def build_rows(filters):
    with app_context():
        query = Order.query
        if filters.get("manager"):
            query = query.join(Manager).filter(Manager.code == filters["manager"])
        if filters.get("status"):
            query = query.filter(Order.status == filters["status"])
        if filters.get("start_date"):
            query = query.filter(Order.order_date >= filters["start_date"])
        if filters.get("end_date"):
            query = query.filter(Order.order_date <= filters["end_date"])

        orders = query.order_by(Order.order_date.desc()).all()

        rows = []
        for order in orders:
            for item in order.items:
                notes = item.notes or ""
                rows.append(
                    {
                        "manager": order.manager.code if order.manager else "",
                        "barcode": extract_field(notes, "바코드"),
                        "order_date": order.order_date.strftime("%Y%m%d")
                        if order.order_date
                        else "",
                        "user_id": order.buyer.user_id if order.buyer else "",
                        "order_no": order.order_no,
                        "buyer_name": order.buyer.name if order.buyer else "",
                        "consignor_name": order.consignor.name if order.consignor else "",
                        "category": extract_field(notes, "브랜드"),
                        "product_name": item.product_name,
                        "color": item.color,
                        "size": extract_field(notes, "사이즈"),
                        "quantity": item.quantity,
                        "options": extract_field(notes, "상가"),
                        "wholesale_price": extract_field(notes, "도매가"),
                        "supplier": extract_field(notes, "미송"),
                        "notes": extract_field(notes, "비고"),
                        "recipient_name": extract_field(notes, "이름"),
                        "phone": extract_field(notes, "전화번호")
                        or (order.buyer.phone if order.buyer else ""),
                        "address": extract_field(notes, "주소"),
                        "buyer_user_id": extract_field(notes, "구매아이디")
                        or (order.buyer.user_id if order.buyer else ""),
                        "delivery_msg": extract_field(notes, "배송메세지"),
                        "code": extract_field(notes, "코드"),
                        "status": item.status or order.status or "입고대기",
                        "status_history": item.status_history
                        or item.status
                        or order.status
                        or "입고대기",
                        "change_log": item.change_log or "",
                    }
                )

        search = (filters.get("search") or "").strip().lower()
        if search:
            filtered = []
            for row in rows:
                if any(search in str(value).lower() for value in row.values()):
                    filtered.append(row)
            rows = filtered

        return rows


def build_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "주문목록"

    headers = [
        "알파벳",
        "미등록주문",
        "주문일",
        "아이디",
        "고유번호",
        "주문자명",
        "위탁자명",
        "브랜드",
        "상품명",
        "색상",
        "사이즈",
        "수량",
        "상가",
        "도매가",
        "미송",
        "비고",
        "이름",
        "전화번호",
        "주소",
        "아이디",
        "배송메세지",
        "코드",
        "상품상태",
    ]
    ws.append(headers)

    color_fills = {
        "입고": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
        "미송": PatternFill(start_color="FF00FFFF", end_color="FF00FFFF", fill_type="solid"),
        "품절": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
        "교환": PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
        "환불": PatternFill(start_color="FFE6B8B7", end_color="FFE6B8B7", fill_type="solid"),
        "택배비": PatternFill(start_color="FFBFBFBF", end_color="FFBFBFBF", fill_type="solid"),
    }

    with app_context():
        orders = Order.query.order_by(Order.order_date.desc()).all()

        for order in orders:
            for item in order.items:
                notes = item.notes or ""

                def extract_field_from_notes(field_name):
                    return extract_field(notes, field_name)

                row = [
                    order.manager.code if order.manager else "",
                    extract_field_from_notes("바코드"),
                    order.order_date.strftime("%Y%m%d") if order.order_date else "",
                    order.buyer.user_id if order.buyer else "",
                    order.order_no,
                    order.buyer.name if order.buyer else "",
                    order.consignor.name if order.consignor else "",
                    extract_field_from_notes("브랜드"),
                    item.product_name,
                    item.color,
                    extract_field_from_notes("사이즈"),
                    item.quantity,
                    extract_field_from_notes("상가"),
                    extract_field_from_notes("도매가"),
                    extract_field_from_notes("미송"),
                    extract_field_from_notes("비고"),
                    extract_field_from_notes("이름"),
                    extract_field_from_notes("전화번호")
                    or (order.buyer.phone if order.buyer else ""),
                    extract_field_from_notes("주소"),
                    extract_field_from_notes("구매아이디")
                    or (order.buyer.user_id if order.buyer else ""),
                    extract_field_from_notes("배송메세지"),
                    extract_field_from_notes("코드"),
                    item.status,
                ]
                ws.append(row)

                row_idx = ws.max_row
                if item.status in color_fills:
                    ws.cell(row_idx, 9).fill = color_fills[item.status]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_grid(rows, visible_fields):
    if not rows:
        st.info("표시할 데이터가 없습니다.")
        return

    df = pd.DataFrame(rows)
    df = df[visible_fields]

    status_cell_style = JsCode(
        """
        function(params) {
            const colors = {
                '입고대기': '#ffffff',
                '입고': '#fff3cd',
                '미송': '#d1ecf1',
                '품절': '#f8d7da',
                '교환': '#fff3e0',
                '환불': '#fce4ec',
                '택배비': '#e0e0e0'
            };
            const status = params.value || '';
            return { backgroundColor: colors[status] || '#ffffff' };
        }
        """
    )

    product_cell_style = JsCode(
        """
        function(params) {
            const colors = {
                '입고대기': '#ffffff',
                '입고': '#fff3cd',
                '미송': '#d1ecf1',
                '품절': '#f8d7da',
                '교환': '#fff3e0',
                '환불': '#fce4ec',
                '택배비': '#e0e0e0'
            };
            const status = params.data ? params.data.status : '';
            return { backgroundColor: colors[status] || '#ffffff' };
        }
        """
    )

    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(sortable=True, filter=True, resizable=True)
    gb.configure_grid_options(rowSelection="multiple", enableRangeSelection=True)

    gb.configure_column("product_name", header_name="상품명", cellStyle=product_cell_style, width=250)
    gb.configure_column("status", header_name="상품상태", cellStyle=status_cell_style, width=110)
    gb.configure_column("status_history", header_name="상태이력", width=200)
    gb.configure_column("change_log", header_name="변경내용", width=300, wrapText=True, autoHeight=True)

    header_map = {
        "manager": "알파벳",
        "barcode": "미등록주문",
        "order_date": "주문일",
        "user_id": "아이디",
        "order_no": "고유번호",
        "buyer_name": "주문자명",
        "consignor_name": "위탁자명",
        "category": "브랜드",
        "product_name": "상품명",
        "color": "색상",
        "size": "사이즈",
        "quantity": "수량",
        "options": "상가",
        "wholesale_price": "도매가",
        "supplier": "미송",
        "notes": "비고",
        "recipient_name": "이름",
        "phone": "전화번호",
        "address": "주소",
        "buyer_user_id": "아이디",
        "delivery_msg": "배송메세지",
        "code": "코드",
        "status": "상품상태",
        "status_history": "상태이력",
        "change_log": "변경내용",
    }

    for field, header in header_map.items():
        if field not in ["product_name", "status", "status_history", "change_log"]:
            gb.configure_column(field, header_name=header)

    grid_options = gb.build()
    AgGrid(
        df,
        gridOptions=grid_options,
        height=600,
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
    )


def main():
    st.set_page_config(page_title="주문 관리 시스템 (Streamlit)", layout="wide")
    st.title("주문 관리 시스템 (Streamlit)")

    os.makedirs(UPLOAD_DIR, exist_ok=True)

    with app_context():
        managers = [m.code for m in Manager.query.filter_by(is_active=True).all()]

    with st.sidebar:
        st.subheader("필터")
        manager = st.selectbox("담당자", ["전체"] + managers)
        status = st.selectbox(
            "상품상태",
            ["전체", "입고대기", "입고", "미송", "품절", "교환", "환불", "택배비"],
        )
        use_date_filter = st.checkbox("날짜 필터 사용", value=False)
        start_date = None
        end_date = None
        if use_date_filter:
            start_date = st.date_input("시작일", value=datetime.today())
            end_date = st.date_input("종료일", value=datetime.today())
        search = st.text_input("검색", placeholder="주문자명/고유번호/상품명 등")

    filters = {
        "manager": None if manager == "전체" else manager,
        "status": None if status == "전체" else status,
        "start_date": start_date,
        "end_date": end_date,
        "search": search,
    }

    with st.spinner("데이터를 불러오는 중..."):
        rows = build_rows(filters)

    total_products = len(rows)
    total_quantity = sum(row.get("quantity") or 0 for row in rows)

    st.caption(f"총 상품 수: {total_products} | 총 수량: {total_quantity}")

    column_labels = [
        ("manager", "알파벳"),
        ("barcode", "미등록주문"),
        ("order_date", "주문일"),
        ("user_id", "아이디(주문)"),
        ("order_no", "고유번호"),
        ("buyer_name", "주문자명"),
        ("consignor_name", "위탁자명"),
        ("category", "브랜드"),
        ("product_name", "상품명"),
        ("color", "색상"),
        ("size", "사이즈"),
        ("quantity", "수량"),
        ("options", "상가"),
        ("wholesale_price", "도매가"),
        ("supplier", "미송"),
        ("notes", "비고"),
        ("recipient_name", "이름"),
        ("phone", "전화번호"),
        ("address", "주소"),
        ("buyer_user_id", "아이디(구매)"),
        ("delivery_msg", "배송메세지"),
        ("code", "코드"),
        ("status", "상품상태"),
        ("status_history", "상태이력"),
        ("change_log", "변경내용"),
    ]

    default_fields = [field for field, _label in column_labels]
    label_map = {label: field for field, label in column_labels}
    label_options = [label for _field, label in column_labels]

    st.markdown("### 엑셀 뷰")
    selected_labels = st.multiselect(
        "표시할 컬럼 선택",
        label_options,
        default=label_options,
    )
    visible_fields = [label_map[label] for label in selected_labels] or default_fields

    render_grid(rows, visible_fields)

    st.markdown("### 엑셀 업로드")
    uploaded_file = st.file_uploader("엑셀 파일 업로드", type=["xlsx", "xls"])
    if uploaded_file is not None:
        if st.button("업로드 실행"):
            safe_name = secure_filename(uploaded_file.name)
            file_path = os.path.join(UPLOAD_DIR, safe_name)
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            with st.spinner("업로드 처리 중..."):
                result = import_excel_complete(file_path)
            if result.get("success"):
                st.success(
                    f"업로드 성공 - 신규 {result.get('inserted', 0)}건, "
                    f"수정 {result.get('updated', 0)}건"
                )
                output_file = result.get("output_file")
                if output_file:
                    output_path = os.path.join(UPLOAD_DIR, output_file)
                    if os.path.exists(output_path):
                        with open(output_path, "rb") as f:
                            st.download_button(
                                "업로드 결과 엑셀 다운로드",
                                data=f.read(),
                                file_name=output_file,
                            )
            else:
                st.error(result.get("error", "알 수 없는 오류"))

    st.markdown("### 업로드 이력 / 롤백")
    with app_context():
        history_rows = UploadHistory.query.order_by(UploadHistory.upload_date.desc()).limit(20).all()
        history_data = [
            {
                "id": h.id,
                "filename": h.filename,
                "upload_date": h.upload_date.strftime("%Y-%m-%d %H:%M:%S"),
                "status": h.status,
                "rows_processed": h.rows_processed,
                "rows_inserted": h.rows_inserted,
                "rows_updated": h.rows_updated,
            }
            for h in history_rows
        ]

    st.dataframe(pd.DataFrame(history_data), use_container_width=True)

    rollback_options = {
        f"{h['id']} | {h['filename']} | {h['status']} | {h['upload_date']}": h["id"]
        for h in history_data
    }
    if rollback_options:
        selected_history = st.selectbox("롤백 대상 선택", list(rollback_options.keys()))
        if st.button("업로드 롤백 실행"):
            with st.spinner("롤백 처리 중..."):
                result = do_rollback(rollback_options[selected_history])
            if result.get("success"):
                st.success(f"롤백 완료 - 삭제 {result.get('deleted', 0)}건")
            else:
                st.error(result.get("error", "롤백 오류"))

    st.markdown("### 엑셀 다운로드")
    if st.button("현재 데이터 엑셀로 다운로드"):
        excel_bytes = build_export_excel()
        filename = f"order_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button("엑셀 다운로드", data=excel_bytes, file_name=filename)


if __name__ == "__main__":
    main()

