# -*- coding: utf-8 -*-
"""
엑셀 100% 완벽 복사 업로드 시스템
새로운 고유번호 시스템: 주문자+위탁자 조합별 기본 번호 + 날짜별 순번
"""
from openpyxl import load_workbook
from openpyxl.styles.colors import Color, COLOR_INDEX
from datetime import datetime
from app.models import db, Manager, Buyer, Consignor, Order, OrderItem, UploadHistory, BuyerConsignorCounter
import re
from collections import defaultdict

# 색상 매핑
EXACT_RGB_MAP = {
    "FFFFFF00": "입고",
    "FFFF00": "입고",
    "FF00FFFF": "미송",
    "00FFFF": "미송",
    "FFFF0000": "품절",
    "FF0000": "품절",
    "FFFFC000": "교환",
    "FFC000": "교환",
    "FFE6B8B7": "환불",
    "E6B8B7": "환불",
    "FFBFBFBF": "택배비",
    "BFBFBF": "택배비",
}

THEME_PATTERN_MAP = {
    (0, -0.249977): "택배비",
    (5, 0.599994): "환불",
}

DEFAULT_OFFICE_THEME = {
    0: (0, 0, 0),
    1: (255, 255, 255),
    2: (238, 236, 225),
    3: (31, 73, 125),
    4: (79, 129, 189),
    5: (192, 80, 77),
}


def color_to_rgb(color_obj, workbook=None):
    """색상 객체를 RGB로 변환"""
    if not color_obj:
        return None
    
    if color_obj.type == 'rgb':
        rgb = color_obj.rgb
        if len(rgb) == 8:
            return rgb[2:]
        return rgb
    
    elif color_obj.type == 'theme':
        theme_id = color_obj.theme
        tint = color_obj.tint or 0.0
        
        pattern_key = (theme_id, round(tint, 6))
        if pattern_key in THEME_PATTERN_MAP:
            return None
        
        if theme_id in DEFAULT_OFFICE_THEME:
            base_rgb = DEFAULT_OFFICE_THEME[theme_id]
            return '%02x%02x%02x' % base_rgb
    
    elif color_obj.type == 'indexed':
        try:
            rgb = COLOR_INDEX[color_obj.indexed]
            return rgb[2:] if len(rgb) == 8 else rgb
        except:
            pass
    
    return None


def get_cell_color_status(cell, workbook):
    """셀 색상으로 상태 판별"""
    if not cell.fill or not cell.fill.start_color:
        return '입고대기'
    
    color = cell.fill.start_color
    
    # Theme 패턴 직접 매칭
    if color.type == 'theme':
        theme_id = color.theme
        tint = color.tint or 0.0
        pattern_key = (theme_id, round(tint, 6))
        if pattern_key in THEME_PATTERN_MAP:
            return THEME_PATTERN_MAP[pattern_key]
    
    # RGB 변환 및 매핑
    rgb = color_to_rgb(color, workbook)
    if rgb:
        rgb_upper = rgb.upper()
        if rgb_upper in EXACT_RGB_MAP:
            return EXACT_RGB_MAP[rgb_upper]
    
    return '입고대기'


def import_excel_complete(filepath):
    """
    새로운 고유번호 시스템으로 엑셀 완벽 업로드
    1단계: 모든 데이터 읽기 및 그룹화
    2단계: 그룹별 기본 번호 할당
    3단계: 순번 부여 및 저장
    """
    wb = load_workbook(filepath)
    ws = wb.active
    
    # 업로드 이력 생성
    upload_id = _create_upload_history(filepath)
    
    try:
        # 컬럼 매핑
        col_mapping = _get_column_mapping(ws)
        
        # 1단계: 모든 행 읽기
        all_rows = _read_all_rows(ws, col_mapping, wb)
        
        # 2단계: 그룹화 및 번호 부여
        grouped_orders = _group_and_assign_numbers(all_rows)
        
        # 3단계: DB에 저장
        result = _save_to_database(grouped_orders)
        
        # 업로드 이력 업데이트
        _update_upload_history(upload_id, result)
        
        # 엑셀 다운로드 파일 생성
        output_file = _create_output_excel(filepath, upload_id)
        
        return {
            'success': True,
            'inserted': result['inserted'],
            'updated': result['updated'],
            'upload_id': upload_id,
            'errors': result['errors'],
            'output_file': output_file
        }
    
    except Exception as e:
        _update_upload_history(upload_id, {
            'inserted': 0,
            'updated': 0,
            'errors': [f"전체 오류: {str(e)}"]
        })
        return {
            'success': False,
            'error': str(e),
            'upload_id': upload_id
        }


def _create_upload_history(filepath):
    """업로드 이력 생성"""
    import os
    history = UploadHistory(
        filename=os.path.basename(filepath),
        upload_date=datetime.utcnow(),
        status='처리중'
    )
    db.session.add(history)
    db.session.commit()
    return history.id


def _get_column_mapping(ws):
    """컬럼 매핑 가져오기"""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            col_map[val.strip()] = c
    
    def find_column(name):
        return col_map.get(name)
    
    return {
        'manager': find_column('알파벳') or 1,
        'barcode': find_column('미등록주문') or 2,
        'order_date': find_column('주문일') or 3,
        'user_id': 4,
        'order_no': find_column('고유번호') or 5,
        'buyer': find_column('주문자명') or 6,
        'consignor': find_column('위탁자명') or 7,
        'category': find_column('브랜드') or 8,
        'product': find_column('상품명') or 9,
        'color': find_column('색상') or 10,
        'size': find_column('사이즈') or 11,
        'quantity': find_column('수량') or 12,
        'options': find_column('상가') or 13,
        'wholesale_price': find_column('도매가') or 14,
        'supplier': find_column('미송') or 15,
        'notes': find_column('비고') or 16,
        'recipient_name': find_column('이름') or 17,
        'phone': find_column('전화번호') or 18,
        'address': find_column('주소') or 19,
        'buyer_user_id': 20,
        'delivery_msg': find_column('배송메세지') or 21,
        'code': find_column('코드') or 22,
    }


def _read_all_rows(ws, col_mapping, wb):
    """모든 행 읽기"""
    rows = []
    
    for row_idx in range(2, ws.max_row + 1):
        try:
            row_data = _extract_row_data(ws, row_idx, col_mapping, wb)
            if row_data:
                rows.append(row_data)
        except Exception as e:
            print(f"행 {row_idx} 읽기 오류: {e}")
            continue
    
    return rows


def _extract_row_data(ws, row_idx, col_mapping, wb):
    """행 데이터 추출"""
    # 필수 데이터
    buyer_name = ws.cell(row_idx, col_mapping['buyer']).value
    product_name = ws.cell(row_idx, col_mapping['product']).value
    
    if not buyer_name or not product_name:
        return None
    
    # 담당자 코드
    manager_val = ws.cell(row_idx, col_mapping['manager']).value
    manager_code = _extract_manager_code(manager_val)
    
    # 주문일
    order_date_val = ws.cell(row_idx, col_mapping['order_date']).value
    order_date = _parse_date(order_date_val)
    
    # 위탁자
    consignor_name = ws.cell(row_idx, col_mapping['consignor']).value
    
    # 상태 (색상)
    product_cell = ws.cell(row_idx, col_mapping['product'])
    status = get_cell_color_status(product_cell, wb)
    
    # 기타 데이터
    return {
        'row_idx': row_idx,
        'manager_code': manager_code,
        'buyer_name': buyer_name,
        'consignor_name': consignor_name,
        'order_date': order_date,
        'product_name': product_name,
        'status': status,
        'quantity': ws.cell(row_idx, col_mapping['quantity']).value or 1,
        'color': ws.cell(row_idx, col_mapping['color']).value,
        'user_id': ws.cell(row_idx, col_mapping['user_id']).value,
        'buyer_user_id': ws.cell(row_idx, col_mapping['buyer_user_id']).value,
        'phone': ws.cell(row_idx, col_mapping['phone']).value,
        'barcode': ws.cell(row_idx, col_mapping['barcode']).value,
        'category': ws.cell(row_idx, col_mapping['category']).value,
        'size': ws.cell(row_idx, col_mapping['size']).value,
        'options': ws.cell(row_idx, col_mapping['options']).value,
        'wholesale_price': ws.cell(row_idx, col_mapping['wholesale_price']).value,
        'supplier': ws.cell(row_idx, col_mapping['supplier']).value,
        'notes': ws.cell(row_idx, col_mapping['notes']).value,
        'recipient_name': ws.cell(row_idx, col_mapping['recipient_name']).value,
        'address': ws.cell(row_idx, col_mapping['address']).value,
        'delivery_msg': ws.cell(row_idx, col_mapping['delivery_msg']).value,
        'code': ws.cell(row_idx, col_mapping['code']).value,
    }


def _extract_manager_code(manager_val):
    """담당자 코드 추출"""
    if manager_val and isinstance(manager_val, str):
        match = re.match(r'^([A-Za-z]+)', manager_val.strip())
        if match:
            return match.group(1).upper()[:2]
    return 'XX'


def _parse_date(date_val):
    """날짜 파싱"""
    if isinstance(date_val, datetime):
        return date_val.date()
    elif isinstance(date_val, str):
        try:
            return datetime.strptime(date_val, '%Y%m%d').date()
        except:
            return datetime.today().date()
    else:
        return datetime.today().date()


def _group_and_assign_numbers(all_rows):
    """그룹화 및 고유번호 부여"""
    # (buyer_name, consignor_name, order_date, manager_code)로 그룹화
    groups = defaultdict(list)
    
    for row in all_rows:
        key = (
            row['buyer_name'],
            row['consignor_name'],
            row['order_date'],
            row['manager_code']
        )
        groups[key].append(row)
    
    # 각 그룹에 고유번호 부여
    result = []
    for (buyer_name, consignor_name, order_date, manager_code), items in groups.items():
        # 주문자/위탁자 동일 여부
        is_consignment = (buyer_name == consignor_name)
        prefix = '#' if is_consignment else ''
        
        # 총 개수
        total_count = len(items)
        
        # 기본 번호는 나중에 buyer_id, consignor_id로 결정
        for idx, item in enumerate(items, 1):
            item['seq_number'] = idx
            item['total_count'] = total_count
            item['is_consignment'] = is_consignment
            item['prefix'] = prefix
            result.append(item)
    
    return result


def _save_to_database(grouped_orders):
    """DB에 저장"""
    inserted = 0
    updated = 0
    errors = []
    
    # buyer, consignor, manager 캐시
    buyers_cache = {}
    consignors_cache = {}
    managers_cache = {}
    
    for item_data in grouped_orders:
        try:
            # Manager 가져오기
            manager = _get_or_create_manager(
                item_data['manager_code'],
                managers_cache
            )
            
            # Buyer 가져오기
            buyer = _get_or_create_buyer(
                item_data['buyer_name'],
                item_data['user_id'],
                item_data['phone'],
                buyers_cache
            )
            
            # Consignor 가져오기
            consignor = None
            if item_data['consignor_name']:
                consignor = _get_or_create_consignor(
                    item_data['consignor_name'],
                    consignors_cache
                )
            
            # 기본 번호 가져오기
            base_number = BuyerConsignorCounter.get_base_number(
                buyer.id,
                consignor.id if consignor else None,
                manager.code
            )
            
            # 고유번호 생성
            ymd = item_data['order_date'].strftime('%y%m%d')
            order_no = f"{item_data['prefix']}{manager.code}{ymd}-{base_number}({item_data['seq_number']}/{item_data['total_count']})"
            
            # Order 생성 또는 찾기
            order = Order.query.filter_by(order_no=order_no).first()
            
            if not order:
                order = Order(
                    order_no=order_no,
                    manager_id=manager.id,
                    buyer_id=buyer.id,
                    consignor_id=consignor.id if consignor else None,
                    order_date=item_data['order_date'],
                    status=item_data['status']
                )
                db.session.add(order)
                db.session.flush()
                inserted += 1
            else:
                updated += 1
            
            # OrderItem 생성
            _create_order_item(order, item_data)
            
            # 100개마다 커밋
            if (inserted + updated) % 100 == 0:
                db.session.commit()
        
        except Exception as e:
            errors.append(f"행 {item_data['row_idx']}: {str(e)}")
            continue
    
    db.session.commit()
    
    return {
        'inserted': inserted,
        'updated': updated,
        'errors': errors
    }


def _get_or_create_manager(code, cache):
    """Manager 가져오기 또는 생성"""
    if code in cache:
        return cache[code]
    
    manager = Manager.query.filter_by(code=code).first()
    if not manager:
        manager = Manager(code=code, name=f'담당자{code}')
        db.session.add(manager)
        db.session.flush()
    
    cache[code] = manager
    return manager


def _get_or_create_buyer(name, user_id, phone, cache):
    """Buyer 가져오기 또는 생성"""
    if name in cache:
        return cache[name]
    
    buyer = Buyer.query.filter_by(name=name).first()
    if not buyer:
        buyer = Buyer(name=name, user_id=user_id, phone=phone)
        db.session.add(buyer)
        db.session.flush()
    
    cache[name] = buyer
    return buyer


def _get_or_create_consignor(name, cache):
    """Consignor 가져오기 또는 생성"""
    if name in cache:
        return cache[name]
    
    consignor = Consignor.query.filter_by(name=name).first()
    if not consignor:
        consignor = Consignor(name=name)
        db.session.add(consignor)
        db.session.flush()
    
    cache[name] = consignor
    return consignor


def _create_order_item(order, item_data):
    """OrderItem 생성"""
    # notes 구성
    notes_parts = []
    if item_data['barcode']:
        notes_parts.append(f"바코드: {item_data['barcode']}")
    if item_data['category']:
        notes_parts.append(f"브랜드: {item_data['category']}")
    if item_data['size']:
        notes_parts.append(f"사이즈: {item_data['size']}")
    if item_data['options']:
        notes_parts.append(f"상가: {item_data['options']}")
    if item_data['wholesale_price']:
        notes_parts.append(f"도매가: {item_data['wholesale_price']}")
    if item_data['supplier']:
        notes_parts.append(f"미송: {item_data['supplier']}")
    if item_data['notes']:
        notes_parts.append(f"비고: {item_data['notes']}")
    if item_data['recipient_name']:
        notes_parts.append(f"이름: {item_data['recipient_name']}")
    if item_data['phone']:
        notes_parts.append(f"전화번호: {item_data['phone']}")
    if item_data['address']:
        notes_parts.append(f"주소: {item_data['address']}")
    if item_data['buyer_user_id']:
        notes_parts.append(f"구매아이디: {item_data['buyer_user_id']}")
    if item_data['delivery_msg']:
        notes_parts.append(f"배송메세지: {item_data['delivery_msg']}")
    if item_data['code']:
        notes_parts.append(f"코드: {item_data['code']}")
    
    notes = ", ".join(notes_parts)
    
    # 기존 항목 확인
    existing_item = OrderItem.query.filter_by(
        order_id=order.id,
        product_name=item_data['product_name']
    ).first()
    
    if existing_item:
        # 변경 감지 및 업데이트
        _update_existing_item(existing_item, item_data, notes)
    else:
        # 신규 생성
        item = OrderItem(
            order_id=order.id,
            product_name=item_data['product_name'],
            quantity=int(item_data['quantity']) if item_data['quantity'] else 1,
            color=item_data['color'] or '',
            status=item_data['status'],
            notes=notes,
            status_history=item_data['status'],
            change_log=f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}] 신규 등록"
        )
        db.session.add(item)


def _update_existing_item(existing_item, item_data, notes):
    """기존 항목 업데이트 및 변경 감지"""
    changes = []
    
    # 상태 변경
    if existing_item.status != item_data['status']:
        old_status = existing_item.status
        if existing_item.status_history:
            existing_item.status_history = f"{existing_item.status_history} → {item_data['status']}"
        else:
            existing_item.status_history = f"{old_status} → {item_data['status']}"
        changes.append(f"상태: {old_status} → {item_data['status']}")
        existing_item.status = item_data['status']
    
    # 색상 변경
    new_color = item_data['color'] or ''
    if existing_item.color != new_color and new_color:
        changes.append(f"색상: {existing_item.color or '없음'} → {new_color}")
        existing_item.color = new_color
    
    # 수량 변경
    new_quantity = int(item_data['quantity']) if item_data['quantity'] else 1
    if existing_item.quantity != new_quantity:
        changes.append(f"수량: {existing_item.quantity} → {new_quantity}")
        existing_item.quantity = new_quantity
    
    # notes 변경 감지
    if existing_item.notes != notes:
        # 사이즈 변경
        old_size = re.search(r'사이즈:\s*([^,]+)', existing_item.notes or '')
        new_size = re.search(r'사이즈:\s*([^,]+)', notes)
        if old_size and new_size and old_size.group(1) != new_size.group(1):
            changes.append(f"사이즈: {old_size.group(1)} → {new_size.group(1)}")
        
        # 브랜드 변경
        old_brand = re.search(r'브랜드:\s*([^,]+)', existing_item.notes or '')
        new_brand = re.search(r'브랜드:\s*([^,]+)', notes)
        if old_brand and new_brand and old_brand.group(1) != new_brand.group(1):
            changes.append(f"브랜드: {old_brand.group(1)} → {new_brand.group(1)}")
        
        existing_item.notes = notes
    
    # 변경 로그 기록
    if changes:
        change_entry = f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {', '.join(changes)}"
        if existing_item.change_log:
            existing_item.change_log = f"{existing_item.change_log}\n{change_entry}"
        else:
            existing_item.change_log = change_entry
    
    existing_item.updated_at = datetime.utcnow()


def _update_upload_history(upload_id, result):
    """업로드 이력 업데이트"""
    history = UploadHistory.query.get(upload_id)
    if history:
        history.rows_inserted = result.get('inserted', 0)
        history.rows_updated = result.get('updated', 0)
        history.status = '완료'
        if result.get('errors'):
            history.error_message = '\n'.join(result['errors'][:10])
        db.session.commit()


def _create_output_excel(original_filepath, upload_id):
    """출력 엑셀 파일 생성"""
    try:
        from openpyxl import load_workbook
        import os
        
        wb = load_workbook(original_filepath)
        ws = wb.active
        
        # 헤더 찾기
        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val:
                col_map[val.strip()] = c
        
        col_order_no = col_map.get('고유번호') or 5
        col_buyer = col_map.get('주문자명') or 6
        col_product = col_map.get('상품명') or 9
        
        # 업로드된 주문 가져오기
        history = UploadHistory.query.get(upload_id)
        if history:
            orders = Order.query.filter(
                Order.created_at >= history.upload_date
            ).all()
            
            # 고유번호 매핑
            order_map = {}
            for order in orders:
                for item in order.items:
                    key = f"{order.buyer.name}_{item.product_name}"
                    order_map[key] = order.order_no
            
            # 고유번호 채우기
            for row_idx in range(2, ws.max_row + 1):
                buyer_name = ws.cell(row_idx, col_buyer).value
                product_name = ws.cell(row_idx, col_product).value
                
                if buyer_name and product_name:
                    key = f"{buyer_name}_{product_name}"
                    if key in order_map:
                        ws.cell(row_idx, col_order_no, order_map[key])
            
            # 파일 저장
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'order_{timestamp}.xlsx'
            output_path = os.path.join('uploads', output_filename)
            wb.save(output_path)
            
            return output_filename
    
    except Exception as e:
        print(f"출력 엑셀 생성 오류: {e}")
        return None


# 기존 rollback_upload 함수는 그대로 유지
def rollback_upload(upload_id):
    """업로드 취소 (롤백)"""
    from app.models import DeletedOrderNumber, ManagerCounter
    
    history = UploadHistory.query.get(upload_id)
    if not history:
        return {'success': False, 'error': '업로드 이력을 찾을 수 없습니다'}
    
    orders = Order.query.filter(
        Order.created_at >= history.upload_date
    ).all()
    
    deleted_count = 0
    deleted_numbers = []
    
    for order in orders:
        deleted_numbers.append(order.order_no)
        db.session.delete(order)
        deleted_count += 1
    
    db.session.commit()
    
    return {
        'success': True,
        'deleted': deleted_count,
        'deleted_numbers': deleted_numbers
    }

