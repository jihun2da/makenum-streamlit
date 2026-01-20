# -*- coding: utf-8 -*-
"""
유틸리티 함수
"""
from datetime import datetime
from openpyxl import load_workbook
from app.models import db, Manager, Buyer, Consignor, Order, OrderItem, ManagerCounter
import re

def generate_order_no(manager_code, order_date, manager_id):
    """
    고유번호 생성: M240106-1 (담당자별 누적 번호)
    1. 먼저 삭제된 번호 풀에서 재사용 가능한 번호를 찾음
    2. 없으면 새로운 순번으로 생성 (담당자별 누적)
    Args:
        manager_code: 담당자 코드 (M, A, J 등)
        order_date: 주문 날짜 (date 객체)
        manager_id: 담당자 ID
    Returns:
        str: 고유번호 (예: M240106-1)
    """
    from app.models import DeletedOrderNumber
    import re
    
    # 1단계: 삭제된 번호 풀에서 해당 담당자의 가장 작은 번호 찾기
    deleted_numbers = DeletedOrderNumber.query.filter_by(
        manager_id=manager_id
    ).order_by(DeletedOrderNumber.order_no).all()
    
    # 날짜 형식이 올바른 번호만 재사용 (예: M240106-1)
    for deleted_number in deleted_numbers:
        order_no = deleted_number.order_no
        # 날짜 형식 검증: 알파벳+6자리숫자-숫자 형식인지 확인
        if re.match(r'^[A-Z]+\d{6}-\d+$', order_no):
            # 풀에서 제거
            db.session.delete(deleted_number)
            db.session.commit()
            return order_no
        else:
            # 잘못된 형식의 번호는 풀에서 제거만 함 (재사용 안 함)
            db.session.delete(deleted_number)
    
    db.session.commit()
    
    # 2단계: 재사용 가능한 번호가 없으면 새로운 순번 생성 (담당자별 누적)
    seq = ManagerCounter.get_next_seq(manager_id)
    
    # 형식: M240106-1 (날짜는 주문날짜, 번호는 담당자별 누적)
    ymd = order_date.strftime('%y%m%d')
    return f"{manager_code}{ymd}-{seq}"

def parse_excel_file(filepath):
    """
    엑셀 파일을 파싱하여 DB에 저장
    """
    wb = load_workbook(filepath)
    ws = wb.active
    
    # 헤더 매핑
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if isinstance(val, str):
            headers[val.strip().replace(' ', '')] = c
    
    def find_col(*names):
        for name in names:
            for key, col in headers.items():
                if name in key:
                    return col
        return None
    
    col_manager = find_col('담당')
    col_date = find_col('주문일', '주문날짜', '일자')
    col_user_id = find_col('아이디', 'ID')
    col_buyer = find_col('주문자명')
    col_consignor = find_col('위탁자명')
    col_product = find_col('상품명')
    col_qty = find_col('수량')
    col_status = find_col('상품상태')
    
    processed = 0
    inserted = 0
    updated = 0
    errors = []
    
    for row in range(2, ws.max_row + 1):
        try:
            # 데이터 추출
            manager_val = ws.cell(row=row, column=col_manager).value if col_manager else None
            date_val = ws.cell(row=row, column=col_date).value if col_date else None
            user_id = ws.cell(row=row, column=col_user_id).value if col_user_id else None
            buyer_name = ws.cell(row=row, column=col_buyer).value if col_buyer else None
            consignor_name = ws.cell(row=row, column=col_consignor).value if col_consignor else None
            product_name = ws.cell(row=row, column=col_product).value if col_product else None
            qty = ws.cell(row=row, column=col_qty).value if col_qty else 1
            status = ws.cell(row=row, column=col_status).value if col_status else '입고대기'
            
            # 필수 데이터 검증
            if not all([buyer_name, product_name]):
                continue
            
            # 담당자 코드 추출
            manager_code = extract_manager_code(manager_val, buyer_name)
            manager = Manager.query.filter_by(code=manager_code).first()
            if not manager:
                # 기본 담당자 생성
                manager = Manager(code=manager_code, name=f"담당자{manager_code}")
                db.session.add(manager)
                db.session.flush()
            
            # 주문 날짜
            if isinstance(date_val, datetime):
                order_date = date_val.date()
            elif isinstance(date_val, str):
                try:
                    from dateutil import parser
                    order_date = parser.parse(date_val).date()
                except:
                    order_date = datetime.today().date()
            else:
                order_date = datetime.today().date()
            
            # 주문자 조회/생성
            buyer = Buyer.query.filter_by(name=buyer_name, user_id=user_id).first()
            if not buyer:
                buyer = Buyer(name=buyer_name, user_id=user_id)
                db.session.add(buyer)
                db.session.flush()
            
            # 위탁자 조회/생성
            consignor = None
            if consignor_name:
                consignor = Consignor.query.filter_by(name=consignor_name).first()
                if not consignor:
                    consignor = Consignor(name=consignor_name)
                    db.session.add(consignor)
                    db.session.flush()
            
            # 고유번호 생성
            order_no = generate_order_no(manager.code, order_date, manager.id)
            
            # 기존 주문 확인 (같은 날짜, 담당자, 주문자)
            existing_order = Order.query.filter_by(
                manager_id=manager.id,
                buyer_id=buyer.id,
                order_date=order_date
            ).first()
            
            if existing_order:
                # 기존 주문에 항목 추가
                order = existing_order
                updated += 1
            else:
                # 새 주문 생성
                order = Order(
                    order_no=order_no,
                    manager_id=manager.id,
                    buyer_id=buyer.id,
                    consignor_id=consignor.id if consignor else None,
                    order_date=order_date,
                    status=status or '입고대기'
                )
                db.session.add(order)
                db.session.flush()
                inserted += 1
            
            # 상품 항목 추가
            item = OrderItem(
                order_id=order.id,
                product_name=product_name,
                quantity=int(qty) if isinstance(qty, (int, float)) else 1,
                status=status or '입고대기'
            )
            db.session.add(item)
            
            # 총 수량 업데이트
            order.total_quantity = sum(i.quantity for i in order.items) + item.quantity
            
            processed += 1
            
            # 100행마다 커밋
            if processed % 100 == 0:
                db.session.commit()
        
        except Exception as e:
            errors.append(f"행 {row}: {str(e)}")
            continue
    
    # 최종 커밋
    db.session.commit()
    
    return {
        'processed': processed,
        'inserted': inserted,
        'updated': updated,
        'errors': errors
    }

def extract_manager_code(manager_val, buyer_name):
    """담당자 코드 추출 (영문자만)"""
    if manager_val and isinstance(manager_val, str):
        # 영문자만 추출 (한글 제외)
        match = re.match(r'^([A-Za-z]+)', manager_val.strip())
        if match:
            code = match.group(1).upper()
            # 2글자까지만 (AA, AB 등)
            return code[:2] if len(code) > 1 else code
    
    if buyer_name and isinstance(buyer_name, str):
        # 영문자만 추출
        match = re.match(r'^([A-Za-z]+)', buyer_name.strip())
        if match:
            code = match.group(1).upper()
            return code[:2] if len(code) > 1 else code
    
    return 'XX'  # 기본값을 XX로 (한글자보다 명확)

