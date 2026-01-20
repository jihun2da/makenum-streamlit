# -*- coding: utf-8 -*-
"""
엑셀 100% 완벽 복사 업로드 시스템
"""
from openpyxl import load_workbook
from openpyxl.styles.colors import Color, COLOR_INDEX
from datetime import datetime
from app.models import db, Manager, Buyer, Consignor, Order, OrderItem, UploadHistory
import re

# 색상 매핑 (preserve_styles_and_fill_ids.py에서 가져옴)
EXACT_RGB_MAP = {
    "FFFFFF00": "입고",      # 노랑
    "FFFF00": "입고",        # 노랑 (앞에 FF 없는 경우)
    "FF00FFFF": "미송",      # 청록
    "00FFFF": "미송",        # 청록
    "FFFF0000": "품절",      # 빨강
    "FF0000": "품절",        # 빨강
    "FFFFC000": "교환",      # 주황
    "FFC000": "교환",        # 주황
    "FFBFBFBF": "택배비",    # 회색
    "BFBFBF": "택배비",      # 회색
    "FFE6B8B7": "환불",      # 연분홍
    "E6B8B7": "환불",        # 연분홍
}

# Theme 색상 매핑
THEME_PATTERN_MAP = {
    (0, -0.249977): "택배비",  # Theme 0 (lt1)
    (5, 0.599994): "환불",      # Theme 5 (accent2)
}

def get_cell_color_status(cell, wb):
    """셀의 색상을 분석하여 상태 반환"""
    fill = cell.fill
    if not fill or not fill.patternType:
        return "입고대기"
    
    # Theme 색상 확인
    for col in (fill.start_color, fill.fgColor, fill.bgColor):
        if col and col.type == "theme" and col.theme is not None:
            tint = col.tint if col.tint is not None else 0.0
            for (theme_idx, expected_tint), status in THEME_PATTERN_MAP.items():
                if col.theme == theme_idx and abs(tint - expected_tint) < 0.001:
                    return status
    
    # RGB 색상 확인
    for col in (fill.start_color, fill.fgColor, fill.bgColor):
        if col and col.type == "rgb" and col.rgb:
            rgb = col.rgb.upper()
            # ARGB인 경우 (8자리)
            if len(rgb) == 8:
                rgb_key = rgb  # FF00FFFF 형태
                rgb_key_short = rgb[2:]  # 00FFFF 형태
                if rgb_key in EXACT_RGB_MAP:
                    return EXACT_RGB_MAP[rgb_key]
                if rgb_key_short in EXACT_RGB_MAP:
                    return EXACT_RGB_MAP[rgb_key_short]
            # RGB인 경우 (6자리)
            elif len(rgb) == 6:
                if rgb in EXACT_RGB_MAP:
                    return EXACT_RGB_MAP[rgb]
    
    return "입고대기"

def import_excel_complete(filepath, user_info=None):
    """
    엑셀 파일을 100% 그대로 DB에 저장
    Returns:
        dict: {
            'success': bool,
            'inserted': int,
            'updated': int,
            'upload_id': int,  # 롤백용
            'errors': list
        }
    """
    wb = load_workbook(filepath)
    ws = wb.active
    
    # 업로드 이력 생성
    history = UploadHistory(
        filename=filepath.split('\\')[-1].split('/')[-1],
        status='processing'
    )
    db.session.add(history)
    db.session.flush()
    upload_id = history.id
    
    # 컬럼 매핑 (유연한 매칭)
    col_map = {}
    headers_raw = {}
    
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            val_clean = val.strip()
            col_map[val_clean] = c
            headers_raw[c] = val_clean
    
    # 유연한 컬럼 찾기 함수
    def find_column(*possible_names):
        """여러 가능한 이름 중 하나라도 매칭되면 컬럼 번호 반환"""
        for name in possible_names:
            # 정확히 일치
            if name in col_map:
                return col_map[name]
        return None
    
    # 실제 엑셀 헤더에 맞춰 정확하게 매핑
    # 1. 알파벳 2. 미등록주문 3. 주문일 4. 아이디 5. 고유번호 6. 주문자명 7. 위탁자명
    # 8. 브랜드 9. 상품명 10. 색상 11. 사이즈 12. 수량 13. 상가 14. 도매가
    # 15. 미송 16. 비고 17. 이름 18. 전화번호 19. 주소 20. 아이디 21. 배송메세지 22. 코드 23. 상품상태
    
    # 헤더가 정확히 있으면 find_column으로 찾고, 없으면 순서대로 매핑
    col_manager = find_column('알파벳') if '알파벳' in col_map else 1
    col_barcode = find_column('미등록주문') if '미등록주문' in col_map else 2
    col_order_date = find_column('주문일') if '주문일' in col_map else 3
    col_user_id = 4  # 첫 번째 아이디 (고정)
    col_order_no = find_column('고유번호') if '고유번호' in col_map else 5
    col_buyer = find_column('주문자명') if '주문자명' in col_map else 6
    col_consignor = find_column('위탁자명') if '위탁자명' in col_map else 7
    col_category = find_column('브랜드') if '브랜드' in col_map else 8
    col_product = find_column('상품명') if '상품명' in col_map else 9
    col_color = find_column('색상') if '색상' in col_map else 10
    col_size = find_column('사이즈') if '사이즈' in col_map else 11
    col_quantity = find_column('수량') if '수량' in col_map else 12
    col_options = find_column('상가') if '상가' in col_map else 13
    col_wholesale_price = find_column('도매가') if '도매가' in col_map else 14
    col_supplier = find_column('미송') if '미송' in col_map else 15
    col_notes = find_column('비고') if '비고' in col_map else 16
    col_recipient_name = find_column('이름') if '이름' in col_map else 17
    col_phone = find_column('전화번호') if '전화번호' in col_map else 18
    col_address = find_column('주소') if '주소' in col_map else 19
    col_buyer_user_id = 20  # 두 번째 아이디 (고정)
    col_delivery_msg = find_column('배송메세지') if '배송메세지' in col_map else 21
    col_code = find_column('코드') if '코드' in col_map else 22
    
    # 필수 컬럼 확인
    if not col_buyer or not col_product:
        return {
            'success': False,
            'error': f'필수 컬럼 누락 (헤더: {list(col_map.keys())})',
            'inserted': 0,
            'updated': 0
        }
    
    inserted = 0
    updated = 0
    errors = []
    
    # ===== 1단계: 모든 데이터 읽고 그룹화 =====
    from collections import defaultdict
    from app.models import BuyerConsignorCounter
    
    order_groups = defaultdict(list)  # {(buyer_id, consignor_id, date): [row_data, ...]}
    
    # 첫 번째 패스: 모든 행 읽기
    all_rows_data = []
    for row_idx in range(2, ws.max_row + 1):
        try:
            # 모든 컬럼 데이터 직접 추출 (정확한 순서)
            manager_val = ws.cell(row_idx, col_manager).value  # 1. 알파벳
            barcode_val = ws.cell(row_idx, col_barcode).value  # 2. 미등록주문
            order_date_val = ws.cell(row_idx, col_order_date).value  # 3. 주문일
            user_id_val = ws.cell(row_idx, col_user_id).value  # 4. 아이디
            order_no_val = ws.cell(row_idx, col_order_no).value  # 5. 고유번호
            buyer_name = ws.cell(row_idx, col_buyer).value  # 6. 주문자명
            consignor_name = ws.cell(row_idx, col_consignor).value  # 7. 위탁자명
            category_val = ws.cell(row_idx, col_category).value  # 8. 브랜드
            product_name = ws.cell(row_idx, col_product).value  # 9. 상품명
            color_val = ws.cell(row_idx, col_color).value  # 10. 색상
            size_val = ws.cell(row_idx, col_size).value  # 11. 사이즈
            quantity_val = ws.cell(row_idx, col_quantity).value  # 12. 수량
            options_val = ws.cell(row_idx, col_options).value  # 13. 상가
            wholesale_price_val = ws.cell(row_idx, col_wholesale_price).value  # 14. 도매가
            supplier_val = ws.cell(row_idx, col_supplier).value  # 15. 미송
            notes_val = ws.cell(row_idx, col_notes).value  # 16. 비고
            recipient_name_val = ws.cell(row_idx, col_recipient_name).value  # 17. 이름
            phone_val = ws.cell(row_idx, col_phone).value  # 18. 전화번호
            address_val = ws.cell(row_idx, col_address).value  # 19. 주소
            buyer_user_id_val = ws.cell(row_idx, col_buyer_user_id).value  # 20. 아이디
            delivery_msg_val = ws.cell(row_idx, col_delivery_msg).value  # 21. 배송메세지
            code_val = ws.cell(row_idx, col_code).value  # 22. 코드
            
            # 필수 데이터 확인
            if not buyer_name or not product_name:
                continue
            
            # 담당자 처리
            manager_code = str(manager_val).strip().upper() if manager_val else 'XX'
            # 영문자만 추출
            match = re.match(r'^([A-Za-z]+)', manager_code)
            if match:
                manager_code = match.group(1).upper()[:2]
            else:
                manager_code = 'XX'
            
            manager = Manager.query.filter_by(code=manager_code).first()
            if not manager:
                manager = Manager(code=manager_code, name=f'담당자{manager_code}')
                db.session.add(manager)
                db.session.flush()
            
            # 주문일 처리
            if isinstance(order_date_val, datetime):
                order_date = order_date_val.date()
            elif isinstance(order_date_val, str):
                try:
                    order_date = datetime.strptime(order_date_val, '%Y%m%d').date()
                except:
                    order_date = datetime.today().date()
            else:
                order_date = datetime.today().date()
            
            # 주문자 처리
            buyer = Buyer.query.filter_by(name=buyer_name).first()
            if not buyer:
                buyer = Buyer(
                    name=buyer_name,
                    user_id=user_id_val,
                    phone=phone_val
                )
                db.session.add(buyer)
                db.session.flush()
            
            # 위탁자 처리
            consignor = None
            if consignor_name:
                consignor = Consignor.query.filter_by(name=consignor_name).first()
                if not consignor:
                    consignor = Consignor(name=consignor_name)
                    db.session.add(consignor)
                    db.session.flush()
            
            # 색상으로 상태 판별 (상품명 셀)
            product_cell = ws.cell(row_idx, col_product)
            status = get_cell_color_status(product_cell, wb)
            
            # 고유번호 (엑셀에 있으면 사용, 없으면 생성)
            order_no = order_no_val if order_no_val else ''
            if not order_no:
                from app.utils import generate_order_no
                order_no = generate_order_no(manager.code, order_date, manager.id)
            
            # 주문 생성 또는 찾기
            order = Order.query.filter_by(order_no=order_no).first()
            if order:
                updated += 1
            else:
                order = Order(
                    order_no=order_no,
                    manager_id=manager.id,
                    buyer_id=buyer.id,
                    consignor_id=consignor.id if consignor else None,
                    order_date=order_date,
                    status=status,
                    notes=notes_val or ''
                )
                db.session.add(order)
                db.session.flush()
                inserted += 1
            
            # 상품 항목 추가 (모든 필드 정보를 구조화해서 저장)
            item_notes_parts = []
            if barcode_val:
                item_notes_parts.append(f"바코드: {barcode_val}")
            if category_val:
                item_notes_parts.append(f"브랜드: {category_val}")
            if size_val:
                item_notes_parts.append(f"사이즈: {size_val}")
            if options_val:
                item_notes_parts.append(f"상가: {options_val}")
            if wholesale_price_val:
                item_notes_parts.append(f"도매가: {wholesale_price_val}")
            if supplier_val:
                item_notes_parts.append(f"미송: {supplier_val}")
            if notes_val:
                item_notes_parts.append(f"비고: {notes_val}")
            if recipient_name_val:
                item_notes_parts.append(f"이름: {recipient_name_val}")
            if phone_val:
                item_notes_parts.append(f"전화번호: {phone_val}")
            if address_val:
                item_notes_parts.append(f"주소: {address_val}")
            if buyer_user_id_val:
                item_notes_parts.append(f"구매아이디: {buyer_user_id_val}")
            if delivery_msg_val:
                item_notes_parts.append(f"배송메세지: {delivery_msg_val}")
            if code_val:
                item_notes_parts.append(f"코드: {code_val}")
            
            item_notes = ", ".join(item_notes_parts)
            
            # 기존 주문의 경우 기존 상품 항목 찾기 (상품명으로 매칭)
            existing_item = None
            if order and order.id:
                existing_item = OrderItem.query.filter_by(
                    order_id=order.id,
                    product_name=product_name
                ).first()
            
            # 변경 감지 및 로깅
            changes = []
            status_changed = False
            new_quantity = int(quantity_val) if quantity_val else 1
            new_color = color_val or ''
            
            if existing_item:
                # 기존 항목 업데이트
                # 1. 상태 변경 감지
                if existing_item.status != status:
                    old_status = existing_item.status
                    status_changed = True
                    # 상태 이력 업데이트
                    if existing_item.status_history:
                        existing_item.status_history = f"{existing_item.status_history} → {status}"
                    else:
                        existing_item.status_history = f"{old_status} → {status}"
                    changes.append(f"상태: {old_status} → {status}")
                    existing_item.status = status
                
                # 2. 색상 변경 감지
                if existing_item.color != new_color and new_color:
                    changes.append(f"색상: {existing_item.color or '없음'} → {new_color}")
                    existing_item.color = new_color
                
                # 3. 수량 변경 감지
                if existing_item.quantity != new_quantity:
                    changes.append(f"수량: {existing_item.quantity} → {new_quantity}")
                    existing_item.quantity = new_quantity
                
                # 4. notes에서 사이즈, 브랜드 등 변경 감지
                old_notes = existing_item.notes or ''
                if old_notes != item_notes:
                    # 사이즈 변경
                    old_size = re.search(r'사이즈:\s*([^,]+)', old_notes)
                    new_size = re.search(r'사이즈:\s*([^,]+)', item_notes)
                    if old_size and new_size and old_size.group(1) != new_size.group(1):
                        changes.append(f"사이즈: {old_size.group(1)} → {new_size.group(1)}")
                    
                    # 브랜드 변경
                    old_brand = re.search(r'브랜드:\s*([^,]+)', old_notes)
                    new_brand = re.search(r'브랜드:\s*([^,]+)', item_notes)
                    if old_brand and new_brand and old_brand.group(1) != new_brand.group(1):
                        changes.append(f"브랜드: {old_brand.group(1)} → {new_brand.group(1)}")
                    
                    existing_item.notes = item_notes
                
                # 변경 로그 추가
                if changes:
                    change_entry = f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {', '.join(changes)}"
                    if existing_item.change_log:
                        existing_item.change_log = f"{existing_item.change_log}\n{change_entry}"
                    else:
                        existing_item.change_log = change_entry
                
                existing_item.updated_at = datetime.utcnow()
                item = existing_item
            else:
                # 신규 항목 생성
                item = OrderItem(
                    order_id=order.id,
                    product_name=product_name or '',
                    quantity=new_quantity,
                    color=new_color,
                    status=status,
                    notes=item_notes,
                    status_history=status,  # 초기 상태
                    change_log=f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}] 신규 등록"
                )
                db.session.add(item)
            
            # 총 수량 업데이트
            order.total_quantity = db.session.query(
                db.func.sum(OrderItem.quantity)
            ).filter(OrderItem.order_id == order.id).scalar() or 0
            
            # 100행마다 커밋
            if (row_idx - 1) % 100 == 0:
                db.session.commit()
        
        except Exception as e:
            errors.append(f"행 {row_idx}: {str(e)}")
            continue
    
    # 최종 커밋
    db.session.commit()
    
    # 업로드 이력 업데이트
    history.rows_processed = ws.max_row - 1
    history.rows_inserted = inserted
    history.rows_updated = updated
    history.status = '완료'
    if errors:
        history.error_message = '\n'.join(errors[:10])  # 처음 10개만
    db.session.commit()
    
    # 원본 엑셀에 고유번호 추가한 파일 생성
    output_file = None
    try:
        output_file = create_excel_with_order_numbers(filepath, upload_id)
    except Exception as e:
        print(f"엑셀 생성 오류: {e}")
    
    return {
        'success': True,
        'inserted': inserted,
        'updated': updated,
        'upload_id': upload_id,
        'errors': errors,
        'output_file': output_file
    }

def create_excel_with_order_numbers(original_filepath, upload_id):
    """
    원본 엑셀에 생성된 고유번호를 추가하여 새 파일 생성
    모든 원본 데이터는 그대로 유지하고 고유번호 컬럼만 채움
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from datetime import datetime
    import os
    
    # 원본 파일 로드
    wb = load_workbook(original_filepath)
    ws = wb.active
    
    # 헤더 찾기
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            col_map[val.strip()] = c
    
    col_order_no = col_map.get('고유번호') or 5
    col_product = col_map.get('상품명') or 9
    col_buyer = col_map.get('주문자명') or 6
    
    # 업로드 이력에서 생성된 주문들 가져오기
    from app.models import UploadHistory, Order
    history = UploadHistory.query.get(upload_id)
    if not history:
        return None
    
    # 해당 업로드에서 생성된 주문들
    orders = Order.query.filter(
        Order.created_at >= history.upload_date
    ).all()
    
    # 주문자명 + 상품명으로 고유번호 매핑
    order_map = {}
    for order in orders:
        for item in order.items:
            key = f"{order.buyer.name}_{item.product_name}"
            if key not in order_map:
                order_map[key] = order.order_no
    
    # 색상 정의 (상품상태에 따라)
    color_fills = {
        '입고': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
        '미송': PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid'),
        '품절': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),
        '교환': PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid'),
        '환불': PatternFill(start_color='FFE6B8B7', end_color='FFE6B8B7', fill_type='solid'),
        '택배비': PatternFill(start_color='FFBFBFBF', end_color='FFBFBFBF', fill_type='solid'),
    }
    
    # 각 행에 고유번호 채우기
    for row_idx in range(2, ws.max_row + 1):
        buyer_name = ws.cell(row_idx, col_buyer).value
        product_name = ws.cell(row_idx, col_product).value
        
        if buyer_name and product_name:
            key = f"{buyer_name}_{product_name}"
            if key in order_map:
                # 고유번호 채우기
                ws.cell(row_idx, col_order_no, order_map[key])
    
    # 새 파일 저장 (영문 파일명 사용)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'order_{timestamp}.xlsx'
    output_path = os.path.join('uploads', output_filename)
    wb.save(output_path)
    
    return output_filename

def rollback_upload(upload_id):
    """
    업로드 취소 (롤백)
    1. 주문 삭제
    2. ManagerCounter 정확하게 재계산 (남은 주문의 최대 순번 기준)
    3. 삭제된 번호 풀도 함께 비우기 (번호 재사용 방지)
    """
    from app.models import DeletedOrderNumber, ManagerCounter
    import re
    
    history = UploadHistory.query.get(upload_id)
    if not history:
        return {'success': False, 'error': '업로드 이력을 찾을 수 없습니다'}
    
    # 해당 업로드 시간 이후에 생성된 주문 찾기
    orders = Order.query.filter(
        Order.created_at >= history.upload_date
    ).all()
    
    deleted_count = 0
    deleted_numbers = []
    manager_ids = set()
    
    for order in orders:
        deleted_numbers.append(order.order_no)
        manager_ids.add(order.manager_id)
        
        # 주문 삭제
        db.session.delete(order)
        deleted_count += 1
    
    # 삭제된 번호 풀도 모두 비우기 (해당 담당자들의 모든 삭제된 번호)
    for manager_id in manager_ids:
        DeletedOrderNumber.query.filter_by(manager_id=manager_id).delete()
    
    # ManagerCounter 정확하게 재계산 (남아있는 주문들의 최대 순번 찾기)
    for manager_id in manager_ids:
        # 해당 담당자의 남아있는 모든 주문 조회
        remaining_orders = Order.query.filter_by(manager_id=manager_id).all()
        
        if remaining_orders:
            # 고유번호에서 순번 추출하여 최댓값 찾기
            max_seq = 0
            for order in remaining_orders:
                # 고유번호 형식: AA260108-761 → 761 추출
                match = re.search(r'-(\d+)$', order.order_no)
                if match:
                    seq = int(match.group(1))
                    max_seq = max(max_seq, seq)
            
            # 카운터를 최대 순번으로 설정
            counter = ManagerCounter.query.filter_by(manager_id=manager_id).first()
            if counter:
                counter.current_seq = max_seq
        else:
            # 남은 주문이 없으면 카운터를 0으로 리셋
            counter = ManagerCounter.query.filter_by(manager_id=manager_id).first()
            if counter:
                counter.current_seq = 0
    
    db.session.commit()
    
    return {
        'success': True,
        'deleted': deleted_count,
        'deleted_numbers': deleted_numbers
    }

