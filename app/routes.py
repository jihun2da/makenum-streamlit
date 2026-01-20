# -*- coding: utf-8 -*-
"""
API 라우트
"""
from flask import Blueprint, request, jsonify
from datetime import datetime
from app.models import db, Manager, Buyer, Consignor, Order, OrderItem, UploadHistory
from app.utils import generate_order_no, parse_excel_file
from werkzeug.utils import secure_filename
import os

api_bp = Blueprint('api', __name__)

# ============================================================
# 주문 관리 API
# ============================================================

@api_bp.route('/orders', methods=['GET'])
def get_orders():
    """주문 목록 조회"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    # 필터링
    manager_code = request.args.get('manager')
    status = request.args.get('status')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    
    query = Order.query
    
    if manager_code:
        query = query.join(Manager).filter(Manager.code == manager_code)
    if status:
        query = query.filter(Order.status == status)
    if start_date:
        query = query.filter(Order.order_date >= start_date)
    if end_date:
        query = query.filter(Order.order_date <= end_date)
    if search:
        query = query.join(Buyer).filter(Buyer.name.like(f'%{search}%'))
    
    # 페이지네이션
    pagination = query.order_by(Order.order_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'items': [order.to_dict(include_items=True) for order in pagination.items],
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages
    })

@api_bp.route('/orders/<int:order_id>', methods=['GET'])
def get_order(order_id):
    """주문 상세 조회"""
    order = Order.query.get_or_404(order_id)
    return jsonify(order.to_dict(include_items=True))

@api_bp.route('/orders/<string:order_no>', methods=['GET'])
def get_order_by_no(order_no):
    """고유번호로 주문 조회"""
    order = Order.query.filter_by(order_no=order_no).first_or_404()
    return jsonify(order.to_dict(include_items=True))

@api_bp.route('/orders', methods=['POST'])
def create_order():
    """주문 생성"""
    data = request.get_json()
    
    try:
        # 담당자 조회/생성
        manager = Manager.query.filter_by(code=data['manager_code']).first()
        if not manager:
            return jsonify({'error': '담당자를 찾을 수 없습니다'}), 404
        
        # 주문자 조회/생성
        buyer = Buyer.query.filter_by(
            name=data['buyer_name'],
            user_id=data.get('buyer_user_id')
        ).first()
        if not buyer:
            buyer = Buyer(
                name=data['buyer_name'],
                user_id=data.get('buyer_user_id'),
                member_no=data.get('buyer_member_no'),
                phone=data.get('buyer_phone')
            )
            db.session.add(buyer)
        
        # 위탁자 조회/생성 (선택사항)
        consignor = None
        if data.get('consignor_name'):
            consignor = Consignor.query.filter_by(name=data['consignor_name']).first()
            if not consignor:
                consignor = Consignor(name=data['consignor_name'])
                db.session.add(consignor)
        
        db.session.flush()  # ID 생성
        
        # 주문 날짜
        order_date = datetime.fromisoformat(data['order_date']).date()
        
        # 고유번호 생성
        order_no = generate_order_no(manager.code, order_date, manager.id)
        
        # 주문 생성
        order = Order(
            order_no=order_no,
            manager_id=manager.id,
            buyer_id=buyer.id,
            consignor_id=consignor.id if consignor else None,
            order_date=order_date,
            status=data.get('status', '입고대기'),
            notes=data.get('notes')
        )
        db.session.add(order)
        db.session.flush()
        
        # 주문 상품 항목 추가
        total_qty = 0
        for item_data in data.get('items', []):
            item = OrderItem(
                order_id=order.id,
                product_name=item_data['product_name'],
                quantity=item_data.get('quantity', 1),
                color=item_data.get('color'),
                status=item_data.get('status', '입고대기'),
                notes=item_data.get('notes')
            )
            total_qty += item.quantity
            db.session.add(item)
        
        order.total_quantity = total_qty
        db.session.commit()
        
        return jsonify(order.to_dict(include_items=True)), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@api_bp.route('/orders/<int:order_id>', methods=['PUT'])
def update_order(order_id):
    """주문 수정"""
    order = Order.query.get_or_404(order_id)
    data = request.get_json()
    
    try:
        if 'status' in data:
            order.status = data['status']
        if 'notes' in data:
            order.notes = data['notes']
        
        db.session.commit()
        return jsonify(order.to_dict(include_items=True))
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@api_bp.route('/orders/<int:order_id>', methods=['DELETE'])
def delete_order(order_id):
    """주문 삭제 (삭제된 번호는 재사용 풀에 추가)"""
    from app.models import DeletedOrderNumber
    
    order = Order.query.get_or_404(order_id)
    
    try:
        # 삭제된 번호를 풀에 추가
        deleted_num = DeletedOrderNumber(
            order_no=order.order_no,
            manager_id=order.manager_id,
            order_date=order.order_date
        )
        db.session.add(deleted_num)
        
        # 주문 삭제
        db.session.delete(order)
        db.session.commit()
        return '', 204
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# ============================================================
# 엑셀 업로드 API
# ============================================================

@api_bp.route('/upload', methods=['POST'])
def upload_excel():
    """엑셀 파일 업로드 및 처리 (100% 완벽 복사)"""
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '파일이 선택되지 않았습니다'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Excel 파일만 업로드 가능합니다'}), 400
    
    try:
        filename = secure_filename(file.filename)
        filepath = os.path.join('uploads', filename)
        file.save(filepath)
        
        # 새로운 임포터 사용
        from app.excel_importer import import_excel_complete
        result = import_excel_complete(filepath)
        
        if result['success']:
            return jsonify({
                'message': '업로드 성공',
                'upload_id': result['upload_id'],
                'inserted': result['inserted'],
                'updated': result['updated'],
                'errors': result.get('errors', []),
                'output_file': result.get('output_file')
            }), 200
        else:
            return jsonify({'error': result.get('error', '알 수 없는 오류')}), 400
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """업로드 후 생성된 엑셀 파일 다운로드"""
    from flask import send_from_directory, current_app
    try:
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        return send_from_directory(
            upload_folder,
            filename,
            as_attachment=True
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@api_bp.route('/upload/<int:upload_id>/rollback', methods=['POST'])
def rollback_upload(upload_id):
    """업로드 취소 (마지막 업로드 삭제)"""
    try:
        from app.excel_importer import rollback_upload as do_rollback
        result = do_rollback(upload_id)
        
        if result['success']:
            return jsonify({
                'message': '업로드 취소 완료',
                'deleted': result['deleted']
            }), 200
        else:
            return jsonify({'error': result['error']}), 400
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/upload/history', methods=['GET'])
def get_upload_history():
    """업로드 이력 조회"""
    history = UploadHistory.query.order_by(UploadHistory.upload_date.desc()).limit(20).all()
    return jsonify([h.to_dict() for h in history])

@api_bp.route('/export_excel', methods=['GET'])
def export_excel():
    """현재 데이터를 엑셀로 다운로드 (고유번호, 상품상태 포함)"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from flask import send_file
    import io
    from datetime import datetime
    
    # 모든 주문 데이터 가져오기
    orders = Order.query.order_by(Order.order_date.desc()).all()
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "주문목록"
    
    # 헤더 작성 (23개 컬럼)
    headers = [
        '알파벳', '미등록주문', '주문일', '아이디', '고유번호', '주문자명', '위탁자명',
        '브랜드', '상품명', '색상', '사이즈', '수량', '상가', '도매가',
        '미송', '비고', '이름', '전화번호', '주소', '아이디', '배송메세지', '코드', '상품상태'
    ]
    ws.append(headers)
    
    # 색상 정의
    color_fills = {
        '입고': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
        '미송': PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid'),
        '품절': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),
        '교환': PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid'),
        '환불': PatternFill(start_color='FFE6B8B7', end_color='FFE6B8B7', fill_type='solid'),
        '택배비': PatternFill(start_color='FFBFBFBF', end_color='FFBFBFBF', fill_type='solid'),
    }
    
    # 데이터 작성
    for order in orders:
        for item in order.items:
            # notes에서 추가 정보 파싱
            notes = item.notes or ''
            
            def extract_field(field_name):
                import re
                pattern = rf'{field_name}:\s*([^,]+)'
                match = re.search(pattern, notes)
                return match.group(1).strip() if match else ''
            
            barcode = extract_field('바코드')
            category = extract_field('브랜드')
            size = extract_field('사이즈')
            options = extract_field('상가')
            wholesale_price = extract_field('도매가')
            supplier = extract_field('미송')
            note_text = extract_field('비고')
            recipient_name = extract_field('이름')
            phone = extract_field('전화번호') or order.buyer.phone if order.buyer else ''
            address = extract_field('주소')
            buyer_user_id = extract_field('구매아이디') or (order.buyer.user_id if order.buyer else '')
            delivery_msg = extract_field('배송메세지')
            code = extract_field('코드')
            
            row = [
                order.manager.code if order.manager else '',
                barcode,
                order.order_date.strftime('%Y%m%d') if order.order_date else '',
                order.buyer.user_id if order.buyer else '',
                order.order_no,
                order.buyer.name if order.buyer else '',
                order.consignor.name if order.consignor else '',
                category,
                item.product_name,
                item.color,
                size,
                item.quantity,
                options,
                wholesale_price,
                supplier,
                note_text,
                recipient_name,
                phone,
                address,
                buyer_user_id,
                delivery_msg,
                code,
                item.status
            ]
            ws.append(row)
            
            # 상품명 셀에 색상 적용
            row_idx = ws.max_row
            if item.status in color_fills:
                ws.cell(row_idx, 9).fill = color_fills[item.status]  # 9번 컬럼 = 상품명
    
    # 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 파일명 생성
    filename = f'주문목록_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    try:
        # Flask 2.0+ 사용
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except TypeError:
        # Flask 1.x 호환성
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=filename
        )

# ============================================================
# 통계 API
# ============================================================

@api_bp.route('/stats/daily', methods=['GET'])
def get_daily_stats():
    """일일 통계"""
    date = request.args.get('date', datetime.today().date().isoformat())
    
    stats = db.session.query(
        Manager.code,
        Manager.name,
        Order.status,
        db.func.count(Order.id).label('count'),
        db.func.sum(Order.total_quantity).label('total_qty')
    ).join(Manager).filter(
        Order.order_date == date
    ).group_by(Manager.code, Manager.name, Order.status).all()
    
    result = {}
    for stat in stats:
        if stat.code not in result:
            result[stat.code] = {
                'manager_name': stat.name,
                'statuses': {}
            }
        result[stat.code]['statuses'][stat.status] = {
            'count': stat.count,
            'total_qty': stat.total_qty or 0
        }
    
    return jsonify(result)

# ============================================================
# 기준 데이터 API
# ============================================================

@api_bp.route('/managers', methods=['GET'])
def get_managers():
    """담당자 목록"""
    managers = Manager.query.filter_by(is_active=True).all()
    return jsonify([m.to_dict() for m in managers])

@api_bp.route('/managers', methods=['POST'])
def create_manager():
    """담당자 생성"""
    data = request.get_json()
    
    try:
        manager = Manager(
            code=data['code'],
            name=data['name'],
            email=data.get('email'),
            phone=data.get('phone')
        )
        db.session.add(manager)
        db.session.commit()
        return jsonify(manager.to_dict()), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

